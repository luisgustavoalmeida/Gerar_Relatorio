"""
Gera ficheiros Excel RDO e FT a partir do JSON do cliente e do mapa de células.

Um ficheiro RDO por mês (uma folha por dia com registo). Um ficheiro FT por mês.
Os modelos são copiados de `template/`; o mapa editável está em `template/mapa_celulas_excel.json`.
"""

from __future__ import annotations

import json
import re
import shutil
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from rdo_diario.paths import (
    ARQUIVO_MAPA_CELULAS_EXCEL_JSON,
    PASTA_SAIDA_RELATORIOS_EXCEL,
    RAIZ_PROJETO,
)
from rdo_diario.schema import extrair_horarios_do_registro_dia


def _resolver_caminho_relativo_raiz(rel: str) -> Path:
    p = Path(rel)
    return p if p.is_absolute() else (RAIZ_PROJETO / p)


def carregar_mapa_celulas(caminho: Path | None = None) -> dict[str, Any]:
    alvo = caminho or ARQUIVO_MAPA_CELULAS_EXCEL_JSON
    with alvo.open(encoding="utf-8") as f:
        return json.load(f)


def _slug_seguro(texto: str) -> str:
    t = re.sub(r'[<>:"/\\|?*]+', "_", (texto or "").strip())
    t = re.sub(r"\s+", "_", t)
    return t or "sem_nome"


def caminho_pasta_saida_cliente(documento: dict[str, Any], base: Path | None = None) -> Path:
    """Caminho da pasta de saída Excel do cliente (sem criar pastas)."""
    chave = documento.get("chave") or {}
    c = str(chave.get("contratante") or "").strip()
    n = str(chave.get("natureza_servico") or "").strip()
    raiz = base if base is not None else PASTA_SAIDA_RELATORIOS_EXCEL
    return raiz / _slug_seguro(c) / _slug_seguro(n)


def _pasta_cliente_saida(documento: dict[str, Any], base: Path | None = None) -> Path:
    pasta = caminho_pasta_saida_cliente(documento, base)
    pasta.mkdir(parents=True, exist_ok=True)
    return pasta


def remover_saida_relatorios_excel_cliente(documento: dict[str, Any]) -> None:
    """Apaga a pasta de relatórios Excel gerados para o cliente (se existir)."""
    pasta = caminho_pasta_saida_cliente(documento)
    if pasta.is_dir():
        shutil.rmtree(pasta)
    pasta_contratante = pasta.parent
    if (
        pasta_contratante.is_dir()
        and pasta_contratante != PASTA_SAIDA_RELATORIOS_EXCEL
        and not any(pasta_contratante.iterdir())
    ):
        pasta_contratante.rmdir()


def _hhmm_para_time(texto: str) -> time | None:
    t = (texto or "").strip()
    if not t:
        return None
    t = t.replace(".", ":")
    partes = t.split(":")
    try:
        h = int(partes[0])
        m = int(partes[1]) if len(partes) > 1 else 0
        if not (0 <= h <= 23 and 0 <= m <= 59):
            return None
        return time(h, m)
    except (ValueError, IndexError):
        return None


def _hhmm_para_timedelta(texto: str) -> timedelta | None:
    t = (texto or "").strip()
    if not t:
        return None
    t = t.replace(".", ":")
    partes = t.split(":")
    try:
        h = int(partes[0])
        m = int(partes[1]) if len(partes) > 1 else 0
        if h < 0 or m < 0 or m > 59:
            return None
        return timedelta(hours=h, minutes=m)
    except (ValueError, IndexError):
        return None


def _texto_horarios_ponto(registro: dict[str, Any]) -> str:
    h = extrair_horarios_do_registro_dia(registro)
    rotulos = (
        ("Entrada", "ponto_entrada"),
        ("Saída almoço", "ponto_saida_almoco"),
        ("Entrada almoço", "ponto_entrada_almoco"),
        ("Saída", "ponto_saida"),
        ("Desloc. ida", "deslocamento_ida"),
        ("Desloc. volta", "deslocamento_volta"),
    )
    partes: list[str] = []
    for rot, ch in rotulos:
        v = str(h.get(ch, "") or "").strip()
        if v:
            partes.append(f"{rot}: {v}")
    return " | ".join(partes)


def _valor_metrica_aninhada(registro: dict[str, Any], caminho: str) -> Any:
    partes = caminho.split(".")
    cur: Any = registro
    for p in partes:
        if not isinstance(cur, dict):
            return None
        cur = cur.get(p)
    return cur


def _extrair_primeira_linha(texto: str) -> str:
    """Extrai apenas a primeira linha do texto."""
    linhas = str(texto or "").splitlines()
    return linhas[0] if linhas else ""


def _concatenar_registros_primeira_linha(registro: dict[str, Any]) -> str:
    """Concatena a primeira linha dos campos de registro (servico, extra_escopo, ociosidade)."""
    textos: list[str] = []
    for campo in ("registro_servico", "registro_extra_escopo", "registro_ociosidade"):
        valor = str(registro.get(campo) or "")
        primeira_linha = _extrair_primeira_linha(valor)
        if primeira_linha:
            textos.append(primeira_linha)
    return " ".join(textos).strip()


def _escrever_celula(ws: Any, endereco: str, valor: Any, wrap: bool = False) -> None:
    if not endereco:
        return
    c = ws[endereco]
    c.value = valor
    if wrap:
        al = c.alignment.copy() if c.alignment else Alignment()
        c.alignment = Alignment(
            horizontal=al.horizontal,
            vertical=al.vertical or "top",
            wrap_text=True,
        )


def _datas_registros_por_mes(registros: dict[str, Any]) -> dict[tuple[int, int], list[str]]:
    por_mes: dict[tuple[int, int], list[str]] = {}
    for iso in registros:
        try:
            d = date.fromisoformat(str(iso).strip()[:10])
        except ValueError:
            continue
        chave_mes = (d.year, d.month)
        por_mes.setdefault(chave_mes, []).append(str(iso).strip()[:10])
    for ch in por_mes:
        por_mes[ch].sort()
    return por_mes


def _titulo_planilha_dia(iso_data: str) -> str:
    """Nome de folha Excel (máx. 31 caracteres, sem : * ? / \\ [ ])."""
    t = iso_data[:31]
    for a, b in (("/", "-"), (":", "-"), ("\\", "-"), ("?", ""), ("*", ""), ("[", ""), ("]", "")):
        t = t.replace(a, b)
    return t or "Dia"


def _preencher_rdo_mes(
    documento: dict[str, Any],
    mapa: dict[str, Any],
    ano: int,
    mes: int,
    datas_iso: list[str],
    pasta_saida: Path,
) -> Path:
    cfg = mapa.get("rdo") or {}
    tpl = _resolver_caminho_relativo_raiz(str(cfg.get("arquivo_template") or "template/RDO.xlsx"))
    nome_modelo = str(cfg.get("planilha_modelo") or "")
    if not nome_modelo.strip():
        raise ValueError("mapa: rdo.planilha_modelo em falta.")

    wb = load_workbook(tpl)
    if nome_modelo not in wb.sheetnames:
        raise FileNotFoundError(f"Modelo RDO: folha «{nome_modelo}» não existe em {tpl}.")

    ws_modelo = wb[nome_modelo]
    pares: list[tuple[str, Any]] = []
    for iso in datas_iso:
        dup = wb.copy_worksheet(ws_modelo)
        dup.title = _titulo_planilha_dia(iso)
        pares.append((iso, dup))
    wb.remove(ws_modelo)

    cab = documento.get("cabecalho_fixo") or {}
    registros = documento.get("registros_diarios") or {}
    map_cab = cfg.get("cabecalho_fixo") or {}
    map_dia = cfg.get("por_registro_dia") or {}
    cel_obs = str(cfg.get("observacoes_fiscalizacao_dia") or "").strip()
    cel_hor = str(cfg.get("horarios_ponto_detalhe") or "").strip()

    for iso, ws in pares:
        reg = registros.get(iso) or {}
        if not isinstance(reg, dict):
            reg = {}

        for chave_json, endereco in map_cab.items():
            if not endereco:
                continue
            val = cab.get(chave_json)
            if val is None:
                continue
            _escrever_celula(ws, str(endereco), str(val).strip() if val is not None else "")

        for chave_json, endereco in map_dia.items():
            if not endereco:
                continue

            ed = str(endereco)
            if chave_json == "data_relatorio":
                try:
                    d = date.fromisoformat(iso)
                    _escrever_celula(ws, ed, datetime(d.year, d.month, d.day))
                except ValueError:
                    pass
                continue
            # if chave_json == "numero":
            #     num = reg.get("numero")
            #     if num is not None:
            #         _escrever_celula(ws, ed, str(num))
            #     continue
            # if chave_json == "folha":
            #     fol = str(reg.get("folha") or "").strip()
            #     if fol:
            #         _escrever_celula(ws, ed, fol)
            #     continue
            if chave_json in ("registro_servico", "registro_extra_escopo", "registro_ociosidade"):
                txt = str(reg.get(chave_json) or "")
                _escrever_celula(ws, ed, txt, wrap=True)
                continue
            if chave_json in ("tempo_serviço", "tempo_extra_escopo", "tempo_ociosidade"):
                _escrever_celula(ws, ed, str(reg.get(chave_json) or "").strip())
                continue
            if chave_json in ("ponto_entrada", "ponto_saida", "deslocamento_ida", "deslocamento_volta"):
                hor = _hhmm_para_time(str(reg.get(chave_json) or ""))
                if hor is not None:
                    _escrever_celula(ws, ed, hor)
                else:
                    _escrever_celula(ws, ed, str(reg.get(chave_json) or "").strip())
                continue           
            
            val = reg.get(chave_json)
            if val is None:
                continue
            _escrever_celula(ws, ed, str(val) if val is not None else "")

        if cel_obs:
            _escrever_celula(ws, cel_obs, str(cab.get("fiscalizacao") or "").strip(), wrap=True)
        if cel_hor:
            _escrever_celula(ws, cel_hor, _texto_horarios_ponto(reg), wrap=True)

    nome_f = f"RDO_{ano:04d}-{mes:02d}.xlsx"
    destino = pasta_saida / nome_f
    wb.save(destino)
    wb.close()
    return destino


def _preencher_ft_mes(
    documento: dict[str, Any],
    mapa: dict[str, Any],
    ano: int,
    mes: int,
    datas_iso: list[str],
    pasta_saida: Path,
) -> Path:
    cfg = mapa.get("ft") or {}
    tpl = _resolver_caminho_relativo_raiz(str(cfg.get("arquivo_template") or "template/FT.xlsx"))
    folha = str(cfg.get("folha_dados") or "FICHA TEMPO")

    wb = load_workbook(tpl)
    if folha not in wb.sheetnames:
        raise FileNotFoundError(f"Modelo FT: folha «{folha}» não existe em {tpl}.")
    ws = wb[folha]

    cab = documento.get("cabecalho_fixo") or {}
    registros = documento.get("registros_diarios") or {}
    map_cab = cfg.get("cabecalho_fixo") or {}
    colunas = cfg.get("colunas") or {}
    lin_ini = int(cfg.get("primeira_linha_dia") or 10)
    lin_fim = int(cfg.get("ultima_linha_dia") or 40)
    cel_mes = str(cfg.get("data_mes_referencia_celula") or "A8")

    primeiro = date(ano, mes, 1)
    ws[cel_mes].value = datetime(primeiro.year, primeiro.month, primeiro.day)

    conjunto_iso = set(datas_iso)

    def col_of(key: str) -> str:
        return str(colunas.get(key) or "").strip()

    for linha in range(lin_ini, lin_fim + 1):
        for _k, col in colunas.items():
            if col:
                ws[f"{str(col).strip()}{linha}"].value = None

        dia_mes = linha - lin_ini + 1
        try:
            d = date(ano, mes, dia_mes)
        except ValueError:
            continue

        iso = d.isoformat()
        if iso not in conjunto_iso:
            continue

        reg = registros.get(iso) or {}
        if not isinstance(reg, dict):
            continue
        hor = extrair_horarios_do_registro_dia(reg)

        e, f, g, h = (
            col_of("ponto_entrada"),
            col_of("ponto_saida_almoco"),
            col_of("ponto_entrada_almoco"),
            col_of("ponto_saida"),
        )
        for letra, chave in (
            (e, "ponto_entrada"),
            (f, "ponto_saida_almoco"),
            (g, "ponto_entrada_almoco"),
            (h, "ponto_saida"),
        ):
            if letra:
                t = _hhmm_para_time(str(hor.get(chave) or ""))
                ws[f"{letra}{linha}"].value = t if t is not None else str(hor.get(chave) or "").strip()

        for map_key in (
            "metricas_horas.normais_hhmm",
            "metricas_horas.extra_50_hhmm",
            "metricas_horas.extra_100_hhmm",
            "metricas_horas.adicional_noturno_hhmm",
        ):
            letra = col_of(map_key)
            if not letra:
                continue
            txt = _valor_metrica_aninhada(reg, map_key)
            td = _hhmm_para_timedelta(str(txt or ""))
            ws[f"{letra}{linha}"].value = td if td is not None else str(txt or "").strip()

        ocol = col_of("registro_servico")
        if ocol:
            texto_concatenado = _concatenar_registros_primeira_linha(reg)
            ws[f"{ocol}{linha}"].value = texto_concatenado
            c = ws[f"{ocol}{linha}"]
            al = c.alignment.copy() if c.alignment else Alignment()
            c.alignment = Alignment(
                horizontal=al.horizontal, vertical=al.vertical or "top", wrap_text=True
            )

        wcol = col_of("tipo_dia_feriado_coluna_w")
        if wcol:
            tipo = str((_valor_metrica_aninhada(reg, "metricas_horas.tipo_dia") or "")).strip().lower()
            if tipo == "feriado":
                ws[f"{wcol}{linha}"].value = "FERIADO"

    for chave_json, endereco in map_cab.items():
        if not endereco:
            continue
        val = cab.get(chave_json)
        if val is None:
            continue
        _escrever_celula(ws, str(endereco), str(val).strip() if val is not None else "")

    nome_f = f"FT_{ano:04d}-{mes:02d}.xlsx"
    destino = pasta_saida / nome_f
    wb.save(destino)
    wb.close()
    return destino


def gerar_relatorios_excel(
    documento: dict[str, Any],
    caminho_json: Path | None = None,
    mapa: dict[str, Any] | None = None,
    pasta_saida_base: Path | None = None,
) -> list[Path]:
    """
    Gera todos os pares RDO/FT por mês com base nos registos diários do documento.

    Args:
        documento: conteúdo do JSON (cabecalho_fixo, registros_diarios, chave).
        caminho_json: opcional, apenas para mensagens de contexto.
        mapa: mapa de células (se None, lê `template/mapa_celulas_excel.json`).
        pasta_saida_base: substitui a pasta base `saida_relatorios` (útil para testes).

    Returns:
        Lista de ficheiros Excel criados ou sobrescritos.
    """
    _ = caminho_json
    m = mapa if mapa is not None else carregar_mapa_celulas()
    registros = documento.get("registros_diarios") or {}
    if not isinstance(registros, dict) or not registros:
        raise ValueError("Não há registros_diarios no documento.")

    pasta_cliente = _pasta_cliente_saida(documento, pasta_saida_base)

    por_mes = _datas_registros_por_mes(registros)
    if not por_mes:
        raise ValueError("Nenhuma data ISO válida em registros_diarios.")

    gerados: list[Path] = []
    for (ano, mes), datas_iso in sorted(por_mes.items()):
        gerados.append(_preencher_rdo_mes(documento, m, ano, mes, datas_iso, pasta_cliente))
        gerados.append(_preencher_ft_mes(documento, m, ano, mes, datas_iso, pasta_cliente))

    return gerados
